#!/usr/bin/env python
"""Create two compact supplementary-material Excel workbooks.

Run this file from the project folder after the model runs finish:

    python summarize_supplementary_results.py

Expected input folders (override with command-line options if needed):
    results_10fold_elm_elmabc
    results_10fold_baselines
    results_sr_no_cv

Outputs:
    Supplementary_NonSR_Mesh600_700.xlsx
    Supplementary_SR_Mesh600_700.xlsx
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MESH_RE = re.compile(r"(?:^|[^0-9])Mesh(600|700)(?:[^0-9]|$)", re.IGNORECASE)
EXPECTED_NON_SR_METHODS = {"ELM", "ELMABC", "MLR", "RF", "XGBOOST", "SVR", "GWR"}
IDENTITY = [
    "case", "relative_path", "model", "grid_m", "period", "target",
    "target_variable", "predictors", "status",
]
CV_METRICS = [
    "cv_folds_used",
    "cv_oof_r2_standard", "cv_oof_r2_corr", "cv_oof_rmse", "cv_oof_mae",
    "cv_oof_nrmse_mean", "cv_oof_pbias", "cv_oof_nse",
    "cv_fold_r2_standard_mean", "cv_fold_r2_standard_sd",
    "cv_fold_r2_corr_mean", "cv_fold_r2_corr_sd",
    "cv_fold_rmse_mean", "cv_fold_rmse_sd",
    "cv_fold_mae_mean", "cv_fold_mae_sd",
]
TEST_METRICS = [
    "train_r2_standard", "train_r2_corr", "train_rmse", "train_mae",
    "test_n", "test_r2_standard", "test_r2_corr", "test_rmse", "test_mae",
    "test_nrmse_mean", "test_nrmse_range", "test_pbias", "test_nse",
    "test_pearson_r", "test_observed_mean", "test_predicted_mean",
    "generalization_gap_r2_corr", "test_train_rmse_ratio", "overfitting_flag",
]
SR_DETAIL = [
    "sr_engine", "raw_symbolic_expression", "raw_infix_equation",
    "simplified_equation", "calibration_intercept", "calibration_slope",
    "eureqa_weighted_complexity", "raw_program_length", "raw_program_depth",
    "simplified_node_count", "simplified_operator_count", "simplified_depth",
    "selection_validation_rmse", "selection_validation_r2_standard",
    "selection_validation_pbias", "publication_ready", "publication_status",
    "semantic_features_used", "semantic_feature_count",
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--root", type=Path, default=Path.cwd())
    p.add_argument("--elm-folder", default="results_10fold_elm_elmabc")
    p.add_argument("--baseline-folder", default="results_10fold_baselines")
    p.add_argument("--sr-folder", default="results_sr_no_cv")
    p.add_argument("--output-folder", default="supplementary_material")
    p.add_argument("--overwrite", action="store_true")
    return p.parse_args()


def read_table(folder: Path, stem: str, required: bool = True) -> pd.DataFrame:
    csv_path = folder / f"{stem}.csv"
    xlsx_path = folder / f"{stem}.xlsx"
    if csv_path.is_file():
        return pd.read_csv(csv_path, low_memory=False)
    if xlsx_path.is_file():
        return pd.read_excel(xlsx_path)
    if required:
        raise FileNotFoundError(
            f"Could not find {stem}.csv or {stem}.xlsx in:\n  {folder}"
        )
    return pd.DataFrame()


def is_mesh_600_700(row: pd.Series) -> bool:
    grid = pd.to_numeric(pd.Series([row.get("grid_m")]), errors="coerce").iloc[0]
    if pd.notna(grid) and int(round(float(grid))) in (600, 700):
        return True
    text = " ".join(str(row.get(c, "")) for c in ("case", "relative_path"))
    return bool(MESH_RE.search(text))


def clean_results(df: pd.DataFrame, source: str) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    out.insert(0, "result_source", source)
    out = out[out.apply(is_mesh_600_700, axis=1)].copy()
    if "status" in out:
        out = out[out["status"].fillna("OK").astype(str).str.upper().eq("OK")]
    if "grid_m" in out:
        out["grid_m"] = pd.to_numeric(out["grid_m"], errors="coerce").astype("Int64")
    keys = [c for c in ("case", "model", "grid_m", "target") if c in out]
    if keys:
        out = out.drop_duplicates(keys, keep="last")
    return out.reset_index(drop=True)


def selected_columns(df: pd.DataFrame, wanted: list[str]) -> pd.DataFrame:
    cols = [c for c in wanted if c in df.columns]
    return df.loc[:, cols].copy()


def summarize_non_sr(df: pd.DataFrame) -> pd.DataFrame:
    metrics = [
        c for c in (
            "cv_oof_r2_standard", "cv_oof_r2_corr", "cv_oof_rmse",
            "cv_oof_mae", "test_r2_standard", "test_r2_corr",
            "test_rmse", "test_mae", "test_pbias", "test_nse"
        ) if c in df
    ]
    group_cols = [c for c in ("model", "grid_m", "target_variable") if c in df]
    if not group_cols or not metrics:
        return pd.DataFrame()
    summary = df.groupby(group_cols, dropna=False)[metrics].agg(
        ["count", "mean", "std", "median"]
    ).reset_index()
    summary.columns = [
        "_".join(str(x) for x in col if str(x)) if isinstance(col, tuple) else str(col)
        for col in summary.columns
    ]
    return summary


def summarize_sr(df: pd.DataFrame) -> pd.DataFrame:
    metrics = [
        c for c in (
            "test_r2_standard", "test_r2_corr", "test_rmse", "test_mae",
            "test_pbias", "test_nse", "selection_validation_r2_standard",
            "selection_validation_rmse"
        ) if c in df
    ]
    group_cols = [c for c in ("model", "grid_m", "target_variable") if c in df]
    if not group_cols or not metrics:
        return pd.DataFrame()
    summary = df.groupby(group_cols, dropna=False)[metrics].agg(
        ["count", "mean", "std", "median"]
    ).reset_index()
    summary.columns = [
        "_".join(str(x) for x in col if str(x)) if isinstance(col, tuple) else str(col)
        for col in summary.columns
    ]
    return summary


def definitions(sr: bool) -> pd.DataFrame:
    rows = [
        ("Scope", "Only Mesh600* and Mesh700* cases are included."),
        ("Paper inputs", "Columns 7-15: LST, land value (P), and elevation; mean/min/max for each."),
        ("Paper output", "Column 16; case-specific NHWSCC or Annual HWC."),
        ("R2_standard", "1 - SSE/SST; primary predictive coefficient of determination."),
        ("R2_corr", "Squared Pearson correlation; retained for comparison with the paper."),
        ("RMSE", "Root mean squared error; lower is better."),
        ("MAE", "Mean absolute error; lower is better."),
        ("PBIAS", "100 × sum(predicted-observed)/sum(observed); values near zero are preferred."),
        ("NSE", "Nash-Sutcliffe efficiency; 1 is ideal."),
    ]
    rows.append(
        ("Validation", "SR was run without 10-fold CV; holdout and internal SR validation metrics are reported.")
        if sr else
        ("Validation", "All non-SR methods used 10-fold CV on the training partition; OOF metrics summarize CV predictions.")
    )
    return pd.DataFrame(rows, columns=["Item", "Definition"])


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, cells in enumerate(ws.columns, 1):
            sample = list(cells)[:300]
            width = max((len(str(c.value)) if c.value is not None else 0) for c in sample)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 11), 45)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (float, np.floating)):
                    cell.number_format = "0.0000"
                cell.alignment = Alignment(vertical="top", wrap_text=False)
    wb.save(path)


def write_book(path: Path, sheets: list[tuple[str, pd.DataFrame]], overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"Output already exists (use --overwrite): {path}")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets:
            frame.to_excel(writer, index=False, sheet_name=name[:31])
    format_workbook(path)


def main() -> int:
    args = parse_args()
    root = args.root.expanduser().resolve()
    elm_dir = root / args.elm_folder
    baseline_dir = root / args.baseline_folder
    sr_dir = root / args.sr_folder
    output_dir = root / args.output_folder
    output_dir.mkdir(parents=True, exist_ok=True)

    elm = clean_results(read_table(elm_dir, "all_cases_metrics"), "ELM_ELMABC")
    baselines = clean_results(read_table(baseline_dir, "all_cases_metrics"), "Baselines")
    sr = clean_results(read_table(sr_dir, "all_cases_metrics"), "SR")

    non_sr = pd.concat([elm, baselines], ignore_index=True, sort=False)
    found_methods = {
        re.sub(r"[^A-Z0-9]", "", str(value).upper())
        for value in non_sr.get("model", pd.Series(dtype=str)).dropna()
    }
    missing_methods = sorted(EXPECTED_NON_SR_METHODS - found_methods)
    if missing_methods:
        raise ValueError(
            "Incomplete non-SR results. Missing successful method(s): "
            + ", ".join(missing_methods)
            + ". Check results_10fold_baselines/audit.csv; install required "
              "packages (notably mgwr for GWR) and rerun before summarizing."
        )
    non_sr = selected_columns(
        non_sr,
        ["result_source"] + IDENTITY + CV_METRICS + TEST_METRICS,
    )
    if non_sr.empty:
        raise ValueError("No successful Mesh600/Mesh700 non-SR results were found.")

    sr_case = selected_columns(
        sr,
        ["result_source"] + IDENTITY + TEST_METRICS + SR_DETAIL,
    )
    if sr_case.empty:
        raise ValueError("No successful Mesh600/Mesh700 SR results were found.")

    equations = read_table(sr_dir, "SR_Equations", required=False)
    if not equations.empty:
        equations = clean_results(equations, "SR Equations")
        equations = selected_columns(
            equations,
            ["result_source"] + IDENTITY + SR_DETAIL + [
                "train_r2_standard", "test_r2_standard", "test_r2_corr",
                "test_rmse", "test_mae", "split_strategy", "random_state",
            ],
        )

    non_sr_path = output_dir / "Supplementary_NonSR_Mesh600_700.xlsx"
    sr_path = output_dir / "Supplementary_SR_Mesh600_700.xlsx"
    write_book(
        non_sr_path,
        [
            ("Case Metrics", non_sr),
            ("Method Mesh Summary", summarize_non_sr(non_sr)),
            ("Metric Definitions", definitions(sr=False)),
        ],
        args.overwrite,
    )
    sr_sheets = [
        ("Case Metrics", sr_case),
        ("Method Mesh Summary", summarize_sr(sr_case)),
    ]
    if not equations.empty:
        sr_sheets.append(("Equations", equations))
    sr_sheets.append(("Metric Definitions", definitions(sr=True)))
    write_book(sr_path, sr_sheets, args.overwrite)

    print("Supplementary workbooks created successfully:")
    print(f"  Non-SR: {non_sr_path}")
    print(f"  SR:     {sr_path}")
    print(f"  Non-SR case rows: {len(non_sr)}")
    print(f"  SR case rows:     {len(sr_case)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
